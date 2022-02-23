import pyttsx3
import openpyxl as xl

engine = pyttsx3.init()
voice = engine.getProperty('voices')
engine.setProperty('voice', voice[0].id)
rate = engine.getProperty('rate')
engine.setProperty('rate', 175)


def speak(text):
    engine.say(text)
    engine.runAndWait()


wb = xl.load_workbook('../Datasets/Pokemons Dataset.xlsx')
sheet = wb.active
max_rows = sheet.max_row
max_columns = sheet.max_column

result = "There is no data about this pokemon"


def get_pokemon_data(command):
    global result
    output = []
    for i in range(1, max_rows + 1):
        if command == sheet.cell(i + 1, 1).value:
            index = i + 1
            for item in range(0, max_columns):
                output.append(sheet.cell(index, item + 1).value)

            if output[2] is None:
                if output[10] == "Nothing":
                    result = f"{output[0]} is {output[1]} type pokemon," \
                             f"\nit's attack power is {output[3]}," \
                             f"\ndefence power is {output[4]}," \
                             f"\nHP is {output[5]}," \
                             f"\nspecial attack power is {output[6]}," \
                             f"\nspecial defence power is {output[7]}," \
                             f"\nspeed is {output[8]}," \
                             f"\nit's total power is {output[9]}," \
                             f"\n{output[0]} is strong against {output[10]}" \
                             f"\nit is weak against {output[11]} type pokemons."
                else:
                    result = f"{output[0]} is {output[1]} type pokemon," \
                             f"\nit's attack power is {output[3]}," \
                             f"\ndefence power is {output[4]}," \
                             f"\nHP is {output[5]}," \
                             f"\nspecial attack power is {output[6]}," \
                             f"\nspecial defence power is {output[7]}," \
                             f"\nspeed is {output[8]}," \
                             f"\nit's total power is {output[9]}," \
                             f"\n{output[0]} is strong against {output[10]} type pokemons." \
                             f"\nit is weak against {output[11]} type pokemons."
            else:
                if output[10] == "Nothing":
                    result = f"{output[0]} is {output[1]} and {output[2]} type pokemon," \
                             f"\nit's attack power is {output[3]}," \
                             f"\ndefence power is {output[4]}," \
                             f"\nHP is {output[5]}," \
                             f"\nspecial attack power is {output[6]}," \
                             f"\nspecial defence power is {output[7]}," \
                             f"\nspeed is {output[8]}," \
                             f"\nit's total power is {output[9]}," \
                             f"\n{output[0]} is strong against {output[10]}." \
                             f"\nit is weak against {output[11]} type pokemons."
                else:
                    result = f"{output[0]} is {output[1]} and {output[2]} type pokemon," \
                             f"\nit's attack power is {output[3]}," \
                             f"\ndefence power is {output[4]}," \
                             f"\nHP is {output[5]}," \
                             f"\nspecial attack power is {output[6]}," \
                             f"\nspecial defence power is {output[7]}," \
                             f"\nspeed is {output[8]}," \
                             f"\nit's total power is {output[9]}," \
                             f"\n{output[0]} is strong against {output[10]} type pokemons." \
                             f"\nit is weak against {output[11]} type pokemons."

    return result


pokemon = input("Enter any Pokemon: ")
pokemon = pokemon.title()
result = get_pokemon_data(pokemon)
print(result)
speak(result)
