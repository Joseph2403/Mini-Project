import openpyxl as xl

workbook = xl.load_workbook("../Datasets/Students List.xlsx")
sheet = workbook.active
max_row = sheet.max_row
max_column = sheet.max_column

iterate = True


def details(register_no):
    global iterate, output
    for i in range(2, max_row + 2, 1):
        if register_no == str(sheet.cell(i, 3).value):
            password = input("Password:")
            values = []
            if password == sheet.cell(i, 4).value:
                for j in range(2, max_column + 1, 1):
                    values.append(sheet.cell(i, j).value)
                    iterate = False
                output = f"\nName: {values[0]}" \
                         f"\nAge: {values[3]}" \
                         f"\nD.O.B: {values[4]}:" \
                         f"\nBlood Group: {values[5]}" \
                         f"\nBatch: {values[6]}" \
                         f"\nAddress: {values[7]}" \
                         f"\nContact No: {values[8]}" \
                         f"\nQuota: {values[9]}"
                break
            else:
                output = "Incorrect Password!\nTry Again.\n"
                break
        else:
            output = "Invalid Register Number!\nTry Again.\n"
    return output


while iterate:
    result = details(input("Register Number:"))
    print(result)
